# -*- coding: utf-8 -*-
"""
excel_reporting.py
==================
模块作者：ChatGPT (优化版 V3.1)
更新日期：2025-05-15 (模拟日期)

功能概述:
---------
1.  同时管理和写入"项目特定工作簿"与"中央汇总工作簿"。
2.  支持高分辨率PNG图片的插入，通过二进制缓冲方式避免文件句柄问题。
3.  所有图表和数据块的Excel锚点位置集中在传入的`layout_map`字典中。
4.  工作簿的保存操作在`save_all_workbooks()`方法中集中执行。
5.  提供清晰的API接口用于记录不同类型的实验数据和图像。
6.  优化了日志输出，减少冗余调试信息。
7.  改进了EIS图表处理和Excel文件的错误处理机制。
"""
from __future__ import annotations

import io
import shutil
import time
import logging
from pathlib import Path
from typing import Dict, List, Sequence, Tuple, Union, Optional

import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image as PILImage

# 获取日志记录器但不设置基础配置，保留父级记录器设置
log = logging.getLogger(__name__)

# 常量定义
THIN_BORDER_SIDE = Side(style="thin", color="000000")
DOTTED_BORDER_SIDE = Side(style="dotted", color="D9D9D9")
FULL_THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
FULL_DOTTED_BORDER = Border(left=DOTTED_BORDER_SIDE, right=DOTTED_BORDER_SIDE, top=DOTTED_BORDER_SIDE, bottom=DOTTED_BORDER_SIDE)
DEFAULT_FONT = Font(name="Calibri", size=11)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header_cell(cell: openpyxl.cell.cell.Cell) -> None:
    cell.font = HEADER_FONT
    cell.alignment = CENTER_ALIGNMENT
    cell.fill = HEADER_FILL
    cell.border = FULL_THIN_BORDER


def _style_data_cell(cell: openpyxl.cell.cell.Cell, is_numeric: bool = False) -> None:
    cell.font = DEFAULT_FONT
    cell.alignment = CENTER_ALIGNMENT
    cell.border = FULL_DOTTED_BORDER


def _convert_anchor_to_rowcol(anchor_str: str) -> Tuple[int, int]:
    """将Excel锚点转换为行列索引，如'A1'转换为(1, 1)"""
    if not isinstance(anchor_str, str) or not anchor_str:
        raise ValueError(f"无效的锚点字符串: '{anchor_str}'。锚点不能为空。")
    col_letters = ""
    row_digits = ""
    for char_code in anchor_str:
        if 'A' <= char_code.upper() <= 'Z':
            col_letters += char_code.upper()
        elif '0' <= char_code <= '9':
            row_digits += char_code
    if not col_letters or not row_digits:
        raise ValueError(f"无效的锚点格式: '{anchor_str}'。必须包含列字母和行号。")
    try:
        row = int(row_digits)
        col = column_index_from_string(col_letters)
        if row <= 0 or col <= 0:
             raise ValueError("行号和列号必须为正数。")
        return row, col
    except Exception as e:
        raise ValueError(f"从锚点 '{anchor_str}' 解析行列号失败: {e}")


class ExcelReporter:
    DEFAULT_IMAGE_SIZE_PX: Tuple[int, int] = (520, 360) 
    IT_GALLERY_IMAGES_PER_ROW_PROJECT: int = 2
    IT_GALLERY_IMAGES_PER_ROW_CENTRAL: int = 3
    IT_GALLERY_COL_SPACING: int = 7 
    IT_GALLERY_ROW_SPACING: int = 22 

    def __init__(
        self,
        project_name: str,
        project_excel_path: Union[str, Path],
        central_excel_path: Union[str, Path],
        voltages_list: Optional[Sequence[float]],
        positions_list: Optional[Sequence[str]],
        layout_map: Dict[str, str],
    ):
        """
        初始化ExcelReporter
        
        参数:
        ----
        project_name: 项目名称
        project_excel_path: 项目特定的Excel文件路径
        central_excel_path: 中央汇总Excel文件路径
        voltages_list: 电压序列列表
        positions_list: 位置序列列表
        layout_map: 布局映射字典，包含各种图表的锚点位置
        """
        self.project_name: str = project_name
        self.project_workbook_path: Path = Path(project_excel_path)
        self.central_workbook_path: Path = Path(central_excel_path)
        self.layout_map: Dict[str, str] = layout_map
        self.voltages_list_ref = voltages_list if voltages_list else [] # 用于估算summary行位置

        log.info(f"初始化Excel报告生成器: 项目 '{self.project_name}'")

        # 初始化项目特定工作簿
        project_headers = ["电位 (V)", "积累电荷 (C)", "液体输出位置"]
        self.workbook_project, self.worksheet_project = self._prepare_or_create_workbook(
            self.project_workbook_path, sheet_name=self.project_name, headers=project_headers,
        )
        log.info(f"项目工作簿已准备: '{self.project_workbook_path.name}'")

        # 初始化中央汇总工作簿，使用时间戳确保工作表名唯一
        timestamp_str = time.strftime("%Y%m%d_%H%M%S")
        central_sheet_name = f"{self.project_name}_{timestamp_str}"
        central_headers = [
            "电位 (V)", "积累电荷 (C)", "液体输出位置",
            "项目名", "实验时间", "项目Excel文件名",
        ]
        self.workbook_central, self.worksheet_central = self._prepare_or_create_workbook(
            self.central_workbook_path, sheet_name=central_sheet_name, headers=central_headers,
        )
        log.info(f"中央汇总工作簿已准备: '{self.central_workbook_path.name}' (工作表: {central_sheet_name})")

        # 如果提供了电压和位置信息，预填充数据
        if voltages_list and positions_list and len(voltages_list) == len(positions_list):
            log.info(f"预填充电位-位置数据 ({len(voltages_list)}个点)...")
            for r_offset, (voltage, position) in enumerate(zip(voltages_list, positions_list)):
                current_row = r_offset + 2 
                
                # 项目工作簿数据填充
                _style_data_cell(self.worksheet_project.cell(row=current_row, column=1, value=voltage))
                _style_data_cell(self.worksheet_project.cell(row=current_row, column=2, value=""))
                _style_data_cell(self.worksheet_project.cell(row=current_row, column=3, value=position))

                # 中央工作簿数据填充
                _style_data_cell(self.worksheet_central.cell(row=current_row, column=1, value=voltage))
                _style_data_cell(self.worksheet_central.cell(row=current_row, column=2, value=""))
                _style_data_cell(self.worksheet_central.cell(row=current_row, column=3, value=position))
                _style_data_cell(self.worksheet_central.cell(row=current_row, column=4, value=self.project_name))
                _style_data_cell(self.worksheet_central.cell(row=current_row, column=5, value=time.strftime("%Y-%m-%d")))
                _style_data_cell(self.worksheet_central.cell(row=current_row, column=6, value=str(self.project_workbook_path.name)))
        else:
            log.warning("未提供完整的电压-位置数据，Excel表格初始行将为空。")

    def _prepare_or_create_workbook(self, excel_path: Path, sheet_name: str, headers: List[str]) -> Tuple[Workbook, Worksheet]:
        """准备或创建Excel工作簿，处理工作表和表头"""
        workbook: Workbook
        created_new_workbook = False
        
        # 1. 尝试加载现有工作簿或创建新工作簿
        if excel_path.exists():
            try:
                workbook = openpyxl.load_workbook(excel_path)
                log.info(f"已加载现有工作簿: {excel_path}")
            except Exception as e:
                log.warning(f"无法加载工作簿 {excel_path}: {e}。将创建新工作簿。")
                # 尝试备份损坏的文件
                backup_path = excel_path.with_suffix(f".{time.strftime('%Y%m%d%H%M%S')}.bak")
                try:
                    if backup_path.exists(): 
                        backup_path.unlink(missing_ok=True)
                    shutil.move(str(excel_path), str(backup_path))
                    log.info(f"已备份原工作簿至: {backup_path}")
                except Exception as move_err:
                    log.error(f"备份工作簿失败: {move_err}，将直接覆盖。")
                workbook = openpyxl.Workbook()
                created_new_workbook = True
        else:
            log.info(f"工作簿不存在，将创建新的工作簿: {excel_path}")
            workbook = openpyxl.Workbook()
            created_new_workbook = True

        # 2. 处理默认工作表
        if created_new_workbook and workbook.active and workbook.active.title == "Sheet" and sheet_name != "Sheet":
            if len(workbook.sheetnames) == 1: 
                workbook.active.title = sheet_name
            else: 
                try: 
                    workbook.remove(workbook.active) 
                except Exception as e: 
                    log.warning(f"删除默认工作表失败: {e}")

        # 3. 获取或创建目标工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            log.info(f"使用现有工作表: '{sheet_name}'")
            # 检查现有表头是否完整
            header_row_is_empty = all(worksheet.cell(row=1, column=c_idx + 1).value is None 
                                    for c_idx in range(len(headers)))
        else:
            worksheet = workbook.create_sheet(title=sheet_name)
            log.info(f"已创建新工作表: '{sheet_name}'")
            header_row_is_empty = True 
        
        # 4. 设置表头（如果需要）
        if header_row_is_empty:
            log.info(f"设置工作表 '{sheet_name}' 的表头...")
            for col_idx_plus_1, header_text in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx_plus_1, value=header_text)
                _style_header_cell(cell)
                # 调整列宽
                col_letter = get_column_letter(col_idx_plus_1)
                adjusted_width = max(15, len(str(header_text)) + 5 if header_text else 15) 
                adjusted_width = min(adjusted_width, 50) 
                worksheet.column_dimensions[col_letter].width = adjusted_width

        return workbook, worksheet

    def _insert_image_to_worksheet(
        self,
        worksheet: Worksheet,
        image_path: Optional[Path],
        anchor_str: str,
        image_title: str = "",
        image_size_px: Tuple[int, int] = DEFAULT_IMAGE_SIZE_PX,
    ):
        """
        将图像插入工作表的指定位置
        
        参数:
            worksheet: 目标工作表
            image_path: 图像文件路径
            anchor_str: Excel锚点，如"A1"
            image_title: 图像标题
            image_size_px: 图像尺寸(宽,高)，单位像素
        """
        try:
            row_idx, col_idx = _convert_anchor_to_rowcol(anchor_str)
        except ValueError as e:
            log.error(f"插入图像失败：无效的锚点 '{anchor_str}'. 错误: {e}")
            return

        # 如果有标题，添加标题单元格
        if image_title:
            title_cell_row = max(1, row_idx - 1) 
            title_cell = worksheet.cell(row=title_cell_row, column=col_idx, value=image_title)
            title_cell.font = Font(bold=True, size=10)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")

        # 处理图像插入
        if image_path and image_path.is_file() and image_path.stat().st_size > 0:
            try:
                with open(image_path, 'rb') as img_file:
                    img_data = io.BytesIO(img_file.read())
                    pil_img = PILImage.open(img_data)
                    excel_img = OpenpyxlImage(pil_img)
                
                if image_size_px:
                    excel_img.width, excel_img.height = image_size_px
                
                worksheet.add_image(excel_img, anchor_str)
                log.info(f"图像 '{image_path.name}' 已成功添加到工作表 '{worksheet.title}' 的 '{anchor_str}'.")
            except Exception as e:
                log.error(f"无法将图像 '{image_path.name}' ({image_path}) 插入到工作表 '{worksheet.title}' 的 '{anchor_str}': {e}", exc_info=True)
                worksheet.cell(row=row_idx, column=col_idx, value=f"图片加载错误: {image_path.name}").font = Font(color="FF0000", bold=True)
        else:
            missing_reason = ""
            if image_path is None:
                missing_reason = "未提供图像路径 (None)"
            elif not image_path.exists():
                missing_reason = f"图像文件不存在: {image_path.name}"
            elif not image_path.is_file():
                missing_reason = f"路径不是文件: {image_path.name}"
            elif image_path.stat().st_size == 0:
                missing_reason = f"图像文件为空 (0字节): {image_path.name}"
            else:
                missing_reason = f"未知图像路径问题: {image_path}"
            
            log.warning(f"无法在工作表 '{worksheet.title}' 的 '{anchor_str}' 添加图像。原因: {missing_reason}. 将添加占位符文本。")
            worksheet.cell(row=row_idx, column=col_idx, value=f"图片缺失: {missing_reason}").font = Font(color="808080") # Grey text for placeholder


    def record_main_plot(
        self, tag_name: str, png_image_path: Optional[Path],
        project_anchor_key: str, central_anchor_key: str,
        image_size_px: Tuple[int, int] = DEFAULT_IMAGE_SIZE_PX,
    ):
        """
        记录主要图表（CV、LSV、EIS等）到项目和中央Excel
        
        参数:
            tag_name: 图表标签名称
            png_image_path: 图像文件路径
            project_anchor_key: 项目Excel中的锚点键名
            central_anchor_key: 中央Excel中的锚点键名
            image_size_px: 图像尺寸
        """
        log.info(f"记录图表 '{tag_name}'")
        proj_anchor = self.layout_map.get(project_anchor_key)
        cent_anchor = self.layout_map.get(central_anchor_key)
        plot_title = f"{self.project_name} - {tag_name.replace('_', ' ')} 曲线"

        # 处理项目工作簿中的图表
        if proj_anchor:
            self._insert_image_to_worksheet(
                self.worksheet_project, png_image_path, proj_anchor, plot_title, image_size_px
            )
        else:
            log.error(f"项目Excel缺少锚点键: '{project_anchor_key}' for {tag_name}")

        # 处理中央工作簿中的图表
        if cent_anchor:
            self._insert_image_to_worksheet(
                self.worksheet_central, png_image_path, cent_anchor, plot_title, image_size_px
            )
        else:
            log.error(f"中央Excel缺少锚点键: '{central_anchor_key}' for {tag_name}")


    def record_it_data_row(self, row_index: int, voltage_value: float, charge_text: str, position_text: str):
        """记录IT数据行到Excel"""
        actual_excel_row = row_index + 2  # Excel行号从1开始，第1行是表头
        log.info(f"记录IT数据: 电位={voltage_value:.2f}V, 电荷={charge_text}, 位置={position_text}")

        # 同时更新项目和中央工作表
        for ws in [self.worksheet_project, self.worksheet_central]:
            # 检查电位单元格，如果不匹配则更新
            cell_v_check = ws.cell(row=actual_excel_row, column=1).value
            if cell_v_check is None or abs(cell_v_check - voltage_value) > 1e-3:
                _style_data_cell(ws.cell(row=actual_excel_row, column=1, value=voltage_value))
            
            # 更新电荷值
            _style_data_cell(ws.cell(row=actual_excel_row, column=2, value=charge_text))
            # 位置通常已在初始化时设置，此处仅在必要时更新
            # _style_data_cell(ws.cell(row=actual_excel_row, column=3, value=position_text))


    def add_it_plots_gallery(
        self, it_plot_items: List[Dict[str, any]], target_workbook: str,
        anchor_key: str, image_size_px: Tuple[int, int] = DEFAULT_IMAGE_SIZE_PX
    ):
        """添加IT曲线图库到指定的工作簿"""
        if not it_plot_items:
            log.info(f"IT图库项目列表为空，跳过添加")
            return

        # 确定目标工作表和每行图像数量
        if target_workbook == "project":
            ws = self.worksheet_project
            images_per_row = self.IT_GALLERY_IMAGES_PER_ROW_PROJECT
        else:  # "central"
            ws = self.worksheet_central
            images_per_row = self.IT_GALLERY_IMAGES_PER_ROW_CENTRAL
        
        # 获取起始锚点
        start_anchor_str = self.layout_map.get(anchor_key)
        if not start_anchor_str:
            log.error(f"缺少IT图库起始锚点键: '{anchor_key}'")
            return
        
        try: 
            base_row, base_col = _convert_anchor_to_rowcol(start_anchor_str)
        except ValueError as e:
            log.error(f"IT图库起始锚点 '{start_anchor_str}' 无效: {e}")
            return

        log.info(f"添加IT图库到{target_workbook}工作簿，起始于 {start_anchor_str}，每行{images_per_row}张图")
        
        # 按电压排序图像
        sorted_items = sorted(it_plot_items, key=lambda item: item.get("voltage", float('inf')))

        # 添加所有图像
        for item_idx, item_data in enumerate(sorted_items):
            voltage = item_data.get("voltage", "N/A")
            image_path = item_data.get("path")
            
            # 计算图像在网格中的位置
            row_offset = item_idx // images_per_row
            col_offset = item_idx % images_per_row
            current_anchor = f"{get_column_letter(base_col + col_offset * self.IT_GALLERY_COL_SPACING)}{base_row + row_offset * self.IT_GALLERY_ROW_SPACING}"
            
            # 设置图像标题并插入图像
            image_title = f"IT @ {voltage:.2f} V" if isinstance(voltage, (int, float)) else f"IT @ {voltage}"
            self._insert_image_to_worksheet(ws, image_path, current_anchor, image_title, image_size_px)
        
        log.info(f"IT图库添加完成: {len(sorted_items)}张图")


    def add_summary_info(self, additional_info: Optional[Dict[str, str]] = None):
        """添加项目概要信息到Excel工作簿"""
        # 准备基本信息行
        summary_rows_data = [
            ("项目名称", self.project_name), 
            ("报告生成时间", time.strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        # 添加用户提供的额外信息
        if additional_info: 
            summary_rows_data.extend(additional_info.items())

        # 对项目和中央工作簿分别处理
        for ws_type, ws in [("项目", self.worksheet_project), ("中央", self.worksheet_central)]:
            log.info(f"添加总结信息到{ws_type}工作簿...")
            
            # 估算总结信息的起始行
            start_row_summary = ws.max_row + 3  # 在已有内容后留3行空白
            
            # 考虑IT图库的位置，确保总结信息不会覆盖图库
            it_anchor_key = "it_plots_start_anchor" if ws_type == "项目" else "central_it_plots_start_anchor"
            it_items_per_row = self.IT_GALLERY_IMAGES_PER_ROW_PROJECT if ws_type == "项目" else self.IT_GALLERY_IMAGES_PER_ROW_CENTRAL
            
            it_anchor_str = self.layout_map.get(it_anchor_key)
            if it_anchor_str:
                try:
                    it_base_row, _ = _convert_anchor_to_rowcol(it_anchor_str)
                    num_it_items = len(self.voltages_list_ref)
                    num_it_plot_rows = (num_it_items + it_items_per_row - 1) // it_items_per_row
                    estimated_gallery_end_row = it_base_row + num_it_plot_rows * self.IT_GALLERY_ROW_SPACING
                    start_row_summary = max(start_row_summary, estimated_gallery_end_row + 5)
                except ValueError:
                    pass  # 如果锚点解析失败，使用默认行
            
            # 创建标题行
            title_cell = ws.cell(row=start_row_summary, column=1, value="实验信息概要")
            _style_header_cell(title_cell)
            ws.merge_cells(start_row=start_row_summary, start_column=1, end_row=start_row_summary, end_column=4)
            
            # 添加所有信息行
            current_data_row = start_row_summary + 1
            for key_text, value_text in summary_rows_data:
                key_cell = ws.cell(row=current_data_row, column=1, value=key_text)
                value_cell = ws.cell(row=current_data_row, column=2, value=str(value_text))
                key_cell.font = Font(bold=True)
                key_cell.alignment = Alignment(horizontal="left", vertical="top")
                value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.merge_cells(start_row=current_data_row, start_column=2, end_row=current_data_row, end_column=4)
                current_data_row += 1


    def record_it_result(self, voltage: float, charge_or_status: str, output_pos: str, loop_index: int, it_plot_path: Optional[Path] = None):
        """
        记录单次IT测试结果
        
        参数:
            voltage: 电压值
            charge_or_status: 计算的电荷值或状态文本
            output_pos: 输出位置
            loop_index: 循环索引（对应于voltage_list中的索引）
            it_plot_path: IT曲线图像文件路径
        """
        log.info(f"记录IT结果: 电压={voltage:.2f}V, 电荷={charge_or_status}, 位置={output_pos}, 索引={loop_index}")
        
        # 使用现有的record_it_data_row方法记录表格数据
        self.record_it_data_row(
            row_index=loop_index, 
            voltage_value=voltage, 
            charge_text=charge_or_status, 
            position_text=output_pos
        )
        
        # 如果提供了图像路径，则确保它会在后续添加到图库中
        # 图库在实验完成后通过add_it_plots_gallery方法一次性添加
        
        log.info(f"IT结果记录完成: 索引={loop_index}, 电压={voltage:.2f}V")


    def record_cv_results(self, cv_plot_path: Optional[Path] = None):
        """
        记录CV测试结果
        
        参数:
            cv_plot_path: CV曲线图像文件路径
        """
        log.info(f"记录CV结果图像: {cv_plot_path}")
        
        # 记录CV图表到Excel
        if cv_plot_path:
            self.record_main_plot(
                tag_name="CV", 
                png_image_path=cv_plot_path,
                project_anchor_key="cv_plot_anchor", 
                central_anchor_key="central_cv_plot_anchor",
            )
        
        log.info("CV结果记录完成")


    def save_all_workbooks(self):
        """保存所有工作簿到磁盘"""
        log.info("保存Excel工作簿...")
        saved_count = 0
        
        # 保存项目和中央工作簿
        for wb_path, wb_obj, wb_name_str in [
            (self.project_workbook_path, self.workbook_project, "项目"),
            (self.central_workbook_path, self.workbook_central, "中央汇总")
        ]:
            try:
                # 确保目标目录存在
                wb_path.parent.mkdir(parents=True, exist_ok=True)
                
                # 保存工作簿
                wb_obj.save(wb_path)
                log.info(f"已保存{wb_name_str}工作簿: {wb_path}")
                saved_count += 1
            except Exception as e:
                log.error(f"保存{wb_name_str}工作簿失败: {e}")
        
        if saved_count == 2:
            log.info("所有Excel工作簿保存成功")
        else:
            log.warning(f"部分Excel工作簿保存失败: {saved_count}/2 成功")